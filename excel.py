from openpyxl import load_workbook
from openpyxl.styles import Alignment


class ExcelHandler:
    def __init__(self, file_path: str, sheet_names: list[str]):
        self.file_path = file_path
        self.sheet_names = sheet_names
        self.wb = load_workbook(file_path)
        self.ws = self.wb.active

    def change_sheet(self, sheet_name: str):
        self.ws = self.wb[sheet_name]

    def read_column(self, column: str) -> list[str]:
        data = []
        for cell in self.ws[column]:
            if cell.value is None: break
            data.append(cell.value)
        return data[1:]

    def get_data_rows_count(self, column: str) -> int:
        count = 0
        for cell in self.ws[column]:
            if cell.value is None: break
            count += 1
        return count - 1

    def write_column(self, column: str, data: list[str]):
        row_count = len(data)
        for i, cell in enumerate(self.ws[column][1:row_count+1]):
            cell.value = data[i]
        self.wb.save(self.file_path)

    def merge_cells_and_write(self, start_row: int, start_column: str, end_row: int, end_column: str, data: str):
        self.ws.merge_cells(f'{start_column}{start_row}:{end_column}{end_row}')
        self.ws[f'{start_column}{start_row}'].alignment = Alignment(vertical='center', horizontal='center')
        self.ws[f'{start_column}{start_row}'] = data
        self.wb.save(self.file_path)

